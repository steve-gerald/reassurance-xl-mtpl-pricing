"""
===============================================================================
Devoir de réassurance - Tarification d'un traité XL Motor Third Party Liability
===============================================================================

Script 1 / 7 : Chargement et nettoyage des données
--------------------------------------------------

Ce script lit le classeur Excel fourni par la cédante et le transforme
en trois objets pandas exploitables pour la suite des travaux :
  - df_expo   : table d'exposition (primes émises, nombre de risques) par année
  - df_paie   : triangle de paiements cumulés par sinistre (long format)
  - df_prov   : triangle de provisions Outstanding par sinistre (long format)

Hypothèses :
  - Couche 1 : 8 000 000 XS  2 000 000 (priorité 2 M / plafond 10 M)
  - Couche 2 : 20 000 000 XS 10 000 000 (priorité 10 M / plafond 30 M)
  - Date d'effet du traité : 1er janvier 2024
  - Branche : RC corporelle automobile (Motor Third Party Liability)

Auteur : Steve (jeune actuaire candidat Pricing Actuary)
Date    : Mai 2026
===============================================================================
"""

from pathlib import Path
import pandas as pd
import openpyxl

# ----- Paramètres du devoir ---------------------------------------------------
PROJECT_DIR = Path(__file__).resolve().parent.parent
XLSX_PATH   = PROJECT_DIR / "devoir de reassurance1.xlsx"
OUT_DIR     = PROJECT_DIR / "outputs"
OUT_DIR.mkdir(exist_ok=True)

# Programme de réassurance
PRIORITE_1 = 2_000_000      # Couche 1 - priorité (rétention cédante)
PLAFOND_1  = 10_000_000     # Couche 1 - limite supérieure (priorité + portée)
PRIORITE_2 = 10_000_000     # Couche 2 - priorité
PLAFOND_2  = 30_000_000     # Couche 2 - limite supérieure

ANNEE_TRAITE = 2024          # année de souscription du traité
ANNEES_HIST  = list(range(2014, 2024))


def charger_exposition(xlsx_path: Path) -> pd.DataFrame:
    """Lit l'onglet 'Exposure' et renvoie la table année / prime / nb_risques."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Exposure"]
    rows = list(ws.iter_rows(values_only=True))

    # D'après inspection du classeur, les données d'exposition se trouvent
    # entre les lignes 13 et 23 (1-indexé), colonnes B (année), C (prime),
    # E (nombre de risques).  La ligne 22 (=2023) et la ligne 23 (=2024)
    # sont marquées 'Estimate' en colonne D et F.
    data = []
    for i in range(12, 23):       # indices 0-based 12..22 -> lignes 13..23
        r = rows[i]
        annee = r[1]
        prime = r[2]
        nb    = r[4]
        if annee is None or prime is None:
            continue
        data.append({"annee": int(annee), "prime": float(prime),
                     "nb_risques": float(nb)})
    df = pd.DataFrame(data).sort_values("annee").reset_index(drop=True)
    df["prime_moy"] = df["prime"] / df["nb_risques"]
    return df


def charger_triangle(xlsx_path: Path, sheet_name: str, valeur: str) -> pd.DataFrame:
    """
    Lit un onglet contenant un triangle 'par sinistre' :
      colonne A = date du sinistre
      colonnes B..K = valeur cumulée par année de développement 0..9
    et retourne un DataFrame long format :
      id_sinistre, date_surv, annee_surv, dev (0..9), valeur
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    for idx, r in enumerate(rows):
        if r[0] is None:
            continue
        date_surv = r[0]
        annee_surv = date_surv.year if hasattr(date_surv, "year") else None
        if annee_surv is None:
            continue
        for dev in range(10):
            v = r[1 + dev]
            if v is None:
                continue
            records.append({
                "id_sinistre": idx + 1,
                "date_surv": pd.Timestamp(date_surv),
                "annee_surv": annee_surv,
                "dev": dev,
                valeur: float(v),
            })
    return pd.DataFrame(records)


def construire_df_sinistres(df_paie: pd.DataFrame,
                            df_prov: pd.DataFrame) -> pd.DataFrame:
    """
    Construit la base sinistres par sinistre, avec :
      - dernier paiement cumulé observé
      - dernière provision observée
      - incurred = paiement + provision
      - age de développement à l'observation (dev_max)
    """
    # On garde le dev maximum observé par sinistre
    paie_last = (df_paie.sort_values(["id_sinistre", "dev"])
                       .groupby("id_sinistre")
                       .tail(1)
                       .rename(columns={"dev": "dev_paie",
                                        "paiement": "paiement_last"}))
    prov_last = (df_prov.sort_values(["id_sinistre", "dev"])
                       .groupby("id_sinistre")
                       .tail(1)
                       .rename(columns={"dev": "dev_prov",
                                        "provision": "provision_last"}))

    df = paie_last.merge(prov_last[["id_sinistre", "dev_prov",
                                     "provision_last"]],
                         on="id_sinistre", how="inner")
    df["incurred"] = df["paiement_last"] + df["provision_last"]
    return df.sort_values(["annee_surv", "date_surv"]).reset_index(drop=True)


if __name__ == "__main__":
    print("--- Chargement des données du devoir ---")
    df_expo = charger_exposition(XLSX_PATH)
    df_paie = charger_triangle(XLSX_PATH, "cum.Payments", "paiement")
    df_prov = charger_triangle(XLSX_PATH, "Outstanding reserves", "provision")
    df_sin  = construire_df_sinistres(df_paie, df_prov)

    # Sauvegardes au format parquet (compact) + csv (lisible)
    df_expo.to_csv(OUT_DIR / "exposition.csv", index=False)
    df_paie.to_csv(OUT_DIR / "triangle_paiements.csv", index=False)
    df_prov.to_csv(OUT_DIR / "triangle_provisions.csv", index=False)
    df_sin.to_csv(OUT_DIR / "sinistres.csv", index=False)

    print(f"Exposition : {len(df_expo)} années")
    print(df_expo.to_string(index=False))
    print(f"\nNombre de sinistres : {df_sin['id_sinistre'].nunique()}")
    print(f"Nombre d'observations dans le triangle de paiements : {len(df_paie)}")
    print(f"Nombre d'observations dans le triangle de provisions : {len(df_prov)}")

    print("\nIncurred par année de survenance (M€) :")
    g = (df_sin.assign(incurred_M=df_sin["incurred"] / 1e6)
              .groupby("annee_surv")["incurred_M"].agg(["count", "sum", "max"]))
    g.columns = ["nb_sin", "incurred_total_M€", "incurred_max_M€"]
    print(g.round(2).to_string())
