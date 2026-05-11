"""
Script 8 : Export du fichier Excel récapitulatif
------------------------------------------------

Construit un classeur Excel propre avec les feuilles :
  1. Synthese        : résumé exécutif des principaux résultats
  2. Hypotheses      : programme de réassurance, données et paramètres
  3. Exposition      : table d'exposition par année
  4. Triangle paie   : triangle de paiements cumulés agrégé
  5. BC sans clause  : burning cost sans clause de stabilité
  6. BC avec clause  : burning cost avec clause (deux méthodes)
  7. Pareto          : ajustement GPD sur la sévérité
  8. Prime commerciale : construction finale + recommandation

Utilise openpyxl pour conserver formules et formatage.
"""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_DIR = Path(__file__).resolve().parent.parent
OUT_DIR  = PROJECT_DIR / "outputs"
XLSX_OUT = PROJECT_DIR / "outputs" / "Devoir_Reassurance_Rendu_Steve.xlsx"

# -- Styles globaux -----------------------------------------------------------
TITRE_FONT   = Font(name="Arial", size=14, bold=True, color="FFFFFF")
TITRE_FILL   = PatternFill("solid", start_color="305496")
HEADER_FONT  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", start_color="4472C4")
TOTAL_FONT   = Font(name="Arial", size=10, bold=True)
TOTAL_FILL   = PatternFill("solid", start_color="DDEBF7")
BORDER       = Border(left=Side(style="thin", color="B4B4B4"),
                       right=Side(style="thin", color="B4B4B4"),
                       top=Side(style="thin", color="B4B4B4"),
                       bottom=Side(style="thin", color="B4B4B4"))


def set_titre(ws, row: int, texte: str, span: int = 8):
    ws.cell(row=row, column=1, value=texte)
    ws.cell(row=row, column=1).font = TITRE_FONT
    ws.cell(row=row, column=1).fill = TITRE_FILL
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 24


def set_header(ws, row: int, headers: list, start_col: int = 1):
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def autoformat_money(cell, decimals=0):
    cell.number_format = f'#,##0{"." + "0"*decimals if decimals else ""};(#,##0);-'


def autoformat_pct(cell, decimals=3):
    cell.number_format = f'0.{"0"*decimals}%'


# -----------------------------------------------------------------------------
# Construction du classeur
# -----------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Synthese"

# ---- 1. SYNTHESE ------------------------------------------------------------
set_titre(ws, 1, "Synthèse — Tarification du traité XL MTPL 2024")
ws.cell(row=3, column=1, value="Préparé par :").font = Font(bold=True)
ws.cell(row=3, column=2, value="Steve (candidat Pricing Actuary)")
ws.cell(row=4, column=1, value="Date :").font = Font(bold=True)
ws.cell(row=4, column=2, value="Mai 2026")
ws.cell(row=5, column=1, value="Cédante :").font = Font(bold=True)
ws.cell(row=5, column=2, value="Portefeuille Motor Third Party Liability")

# Programme de réassurance
set_titre(ws, 7, "Programme de réassurance proposé")
set_header(ws, 9, ["Couche", "Priorité (€)", "Portée (€)", "Plafond (€)"])
ws.cell(row=10, column=1, value="Couche 1 : 8 XS 2")
ws.cell(row=10, column=2, value=2_000_000)
ws.cell(row=10, column=3, value=8_000_000)
ws.cell(row=10, column=4, value="=B10+C10")
ws.cell(row=11, column=1, value="Couche 2 : 20 XS 10")
ws.cell(row=11, column=2, value=10_000_000)
ws.cell(row=11, column=3, value=20_000_000)
ws.cell(row=11, column=4, value="=B11+C11")
for row in [10, 11]:
    for col in [2, 3, 4]:
        autoformat_money(ws.cell(row=row, column=col))
        ws.cell(row=row, column=col).border = BORDER
    ws.cell(row=row, column=1).border = BORDER

# Comparaison des méthodes
set_titre(ws, 13, "Estimation de la prime pure 2024 — comparaison des méthodes")
set_header(ws, 15, ["Méthode", "Prime pure Couche 1 (€)",
                     "Prime pure Couche 2 (€)", "BC1 (%)", "BC2 (%)"])

synth_sans = pd.read_csv(OUT_DIR / "synthese_sans_clause.csv").iloc[0]
synth_avec = pd.read_csv(OUT_DIR / "synthese_avec_clause.csv")
synth_pareto = pd.read_csv(OUT_DIR / "synthese_pareto.csv").iloc[0]
prime_2024 = 229_690_344.58

methodes = [
    ("BC sans clause de stabilité",
     synth_sans["prime_pure_couche_1"],
     synth_sans["prime_pure_couche_2"],
     synth_sans["bc_couche_1"], synth_sans["bc_couche_2"]),
    ("BC avec clause - méthode A (as-if simple, 1% sup.)",
     synth_avec[synth_avec["methode"]=="A - As-if simple (1 pct sup.)"]["prime_pure_couche_1"].iloc[0],
     synth_avec[synth_avec["methode"]=="A - As-if simple (1 pct sup.)"]["prime_pure_couche_2"].iloc[0],
     synth_avec[synth_avec["methode"]=="A - As-if simple (1 pct sup.)"]["bc_couche_1"].iloc[0],
     synth_avec[synth_avec["methode"]=="A - As-if simple (1 pct sup.)"]["bc_couche_2"].iloc[0]),
    ("BC avec clause - méthode B (par paiement, 1% sup.) [REFERENCE]",
     synth_avec[synth_avec["methode"]=="B - Indexation par paiement (1 pct sup.)"]["prime_pure_couche_1"].iloc[0],
     synth_avec[synth_avec["methode"]=="B - Indexation par paiement (1 pct sup.)"]["prime_pure_couche_2"].iloc[0],
     synth_avec[synth_avec["methode"]=="B - Indexation par paiement (1 pct sup.)"]["bc_couche_1"].iloc[0],
     synth_avec[synth_avec["methode"]=="B - Indexation par paiement (1 pct sup.)"]["bc_couche_2"].iloc[0]),
    ("Modèle Pareto/GPD",
     synth_pareto["prime_pure_1_2024_M"] * 1e6,
     synth_pareto["prime_pure_2_2024_M"] * 1e6,
     synth_pareto["prime_pure_1_2024_M"] * 1e6 / prime_2024,
     synth_pareto["prime_pure_2_2024_M"] * 1e6 / prime_2024),
]

for i, (nom, pp1, pp2, bc1, bc2) in enumerate(methodes, start=16):
    ws.cell(row=i, column=1, value=nom)
    ws.cell(row=i, column=2, value=pp1)
    ws.cell(row=i, column=3, value=pp2)
    ws.cell(row=i, column=4, value=bc1)
    ws.cell(row=i, column=5, value=bc2)
    autoformat_money(ws.cell(row=i, column=2))
    autoformat_money(ws.cell(row=i, column=3))
    autoformat_pct(ws.cell(row=i, column=4), decimals=3)
    autoformat_pct(ws.cell(row=i, column=5), decimals=3)
    for col in range(1, 6):
        ws.cell(row=i, column=col).border = BORDER

# Recommandation finale
prime_comm = pd.read_csv(OUT_DIR / "prime_commerciale_finale.csv").iloc[0]

set_titre(ws, 22, "Recommandation finale — Prime commerciale")
set_header(ws, 24, ["Couche", "Prime pure (€)", "Chargement sécurité (€)",
                     "Prime technique (€)", "Prime commerciale (€)",
                     "Taux de prime (% prime cédante)"])
# Couche 1
ws.cell(row=25, column=1, value="Couche 1 (8 XS 2)")
ws.cell(row=25, column=2, value=prime_comm["prime_pure_couche_1"])
sig1 = prime_comm["sigma_couche_1"]
ws.cell(row=25, column=3, value=prime_comm["k_ecart_type"] * sig1)
ws.cell(row=25, column=4, value="=B25+C25")
ws.cell(row=25, column=5, value=prime_comm["prime_commerciale_couche_1"])
ws.cell(row=25, column=6, value=prime_comm["taux_prime_couche_1_pct"]/100)
# Couche 2
ws.cell(row=26, column=1, value="Couche 2 (20 XS 10)")
ws.cell(row=26, column=2, value=prime_comm["prime_pure_couche_2"])
sig2 = prime_comm["sigma_couche_2"]
ws.cell(row=26, column=3, value=prime_comm["k_ecart_type"] * sig2)
ws.cell(row=26, column=4, value="=B26+C26")
ws.cell(row=26, column=5, value=prime_comm["prime_commerciale_couche_2"])
ws.cell(row=26, column=6, value=prime_comm["taux_prime_couche_2_pct"]/100)

for row in [25, 26]:
    for col in [2, 3, 4, 5]:
        autoformat_money(ws.cell(row=row, column=col))
        ws.cell(row=row, column=col).border = BORDER
    autoformat_pct(ws.cell(row=row, column=6), decimals=3)
    ws.cell(row=row, column=6).border = BORDER
    ws.cell(row=row, column=1).border = BORDER

# Largeurs de colonnes
ws.column_dimensions["A"].width = 52
for col in "BCDEF":
    ws.column_dimensions[col].width = 22

# ---- 2. HYPOTHESES ----------------------------------------------------------
ws2 = wb.create_sheet("Hypotheses")
set_titre(ws2, 1, "Hypothèses & paramètres")

ws2.cell(row=3, column=1, value="Programme XL").font = Font(bold=True)
hyp = [
    ("Branche", "Motor Third Party Liability (RC corporelle auto)"),
    ("Date d'effet du traité", "01/01/2024"),
    ("Priorité Couche 1", 2_000_000),
    ("Portée Couche 1", 8_000_000),
    ("Priorité Couche 2", 10_000_000),
    ("Portée Couche 2", 20_000_000),
    ("Période historique", "2014 - 2023 (10 années)"),
    ("Nombre de sinistres observés", 114),
    ("", ""),
    ("Paramètres de tarification", ""),
    ("Méthode retenue pour prime pure", "Burning Cost avec clause de stabilité, indexation par paiement"),
    ("Indice d'inflation des sinistres", "CPI + 1 % superimposed (paramètre énoncé)"),
    ("Coefficient chargement sécurité (k)", prime_comm["k_ecart_type"]),
    ("Frais de gestion", prime_comm["frais_gestion"]),
    ("Coût du capital", prime_comm["cout_capital"]),
]
for i, (k, v) in enumerate(hyp, start=4):
    ws2.cell(row=i, column=1, value=k).font = Font(bold=(k != ""))
    ws2.cell(row=i, column=2, value=v)
    if isinstance(v, (int, float)) and v < 1:
        autoformat_pct(ws2.cell(row=i, column=2))
    elif isinstance(v, (int, float)):
        autoformat_money(ws2.cell(row=i, column=2))

ws2.column_dimensions["A"].width = 42
ws2.column_dimensions["B"].width = 48

# ---- 3. EXPOSITION ----------------------------------------------------------
ws3 = wb.create_sheet("Exposition")
set_titre(ws3, 1, "Exposition par année", span=5)
df_expo = pd.read_csv(OUT_DIR / "exposition.csv")
set_header(ws3, 3, ["Année", "Prime émise (€)", "Nombre de risques", "Prime moyenne (€)"])
for i, row in enumerate(df_expo.itertuples(index=False), start=4):
    ws3.cell(row=i, column=1, value=int(row.annee))
    ws3.cell(row=i, column=2, value=float(row.prime))
    ws3.cell(row=i, column=3, value=float(row.nb_risques))
    # Prime moyenne calculée comme formule Excel
    ws3.cell(row=i, column=4, value=f"=B{i}/C{i}")
    autoformat_money(ws3.cell(row=i, column=2))
    ws3.cell(row=i, column=3).number_format = "#,##0.00"
    autoformat_money(ws3.cell(row=i, column=4), decimals=2)
    for col in range(1, 5):
        ws3.cell(row=i, column=col).border = BORDER
ws3.column_dimensions["A"].width = 10
for c in "BCD":
    ws3.column_dimensions[c].width = 20

# ---- 4. TRIANGLE PAIEMENTS --------------------------------------------------
ws4 = wb.create_sheet("Triangle_paie")
set_titre(ws4, 1, "Triangle agrégé des paiements cumulés (€)", span=12)
df_paie = pd.read_csv(OUT_DIR / "triangle_paiements.csv")
tri = (df_paie.groupby(["annee_surv", "dev"])["paiement"].sum().unstack("dev"))
set_header(ws4, 3, ["Année surv."] + [f"Dev {j}" for j in tri.columns])
for i, (ann, row) in enumerate(tri.iterrows(), start=4):
    ws4.cell(row=i, column=1, value=int(ann))
    ws4.cell(row=i, column=1).font = Font(bold=True)
    for j, v in enumerate(row):
        c = ws4.cell(row=i, column=2 + j, value=float(v) if pd.notna(v) else None)
        autoformat_money(c)
        c.border = BORDER
    ws4.cell(row=i, column=1).border = BORDER

ws4.column_dimensions["A"].width = 12
for j in range(len(tri.columns)):
    ws4.column_dimensions[get_column_letter(2+j)].width = 16

# Bloc Chain-Ladder
cl_facteurs = pd.read_csv(OUT_DIR / "chain_ladder_facteurs.csv")
cl_comp = pd.read_csv(OUT_DIR / "chain_ladder_comparaison.csv")

ws4.cell(row=18, column=1, value="Facteurs de développement (link ratios)").font = Font(bold=True, size=12)
set_header(ws4, 20, ["Transition dev"] + [f"{r['dev']}→{r['dev']+1}" for _, r in cl_facteurs.iterrows()])
ws4.cell(row=21, column=1, value="Facteur")
for j, r in enumerate(cl_facteurs.itertuples(index=False)):
    c = ws4.cell(row=21, column=2 + j, value=float(r.facteur))
    c.number_format = "0.0000"
    c.border = BORDER

ws4.cell(row=24, column=1, value="Comparaison ultime CL vs incurred observé").font = Font(bold=True, size=12)
set_header(ws4, 26, ["Année surv.", "Ultime CL (€)", "Incurred observé (€)", "Écart (%)"])
for i, r in enumerate(cl_comp.itertuples(index=False), start=27):
    ws4.cell(row=i, column=1, value=int(r.annee_surv))
    ws4.cell(row=i, column=2, value=float(r.ultime_CL))
    ws4.cell(row=i, column=3, value=float(r.incurred_observe))
    ws4.cell(row=i, column=4, value=f"=(B{i}-C{i})/C{i}")
    autoformat_money(ws4.cell(row=i, column=2))
    autoformat_money(ws4.cell(row=i, column=3))
    autoformat_pct(ws4.cell(row=i, column=4), decimals=2)
    for col in range(1, 5):
        ws4.cell(row=i, column=col).border = BORDER

# ---- 5. BC SANS CLAUSE ------------------------------------------------------
ws5 = wb.create_sheet("BC_sans_clause")
set_titre(ws5, 1, "Burning Cost SANS clause de stabilité", span=7)
df_bc1 = pd.read_csv(OUT_DIR / "bc_sans_clause.csv")
set_header(ws5, 3, ["Année surv.", "Nb sin.", "Coût Couche 1 (€)",
                    "Coût Couche 2 (€)", "Prime émise (€)",
                    "BC Couche 1", "BC Couche 2"])
for i, r in enumerate(df_bc1.itertuples(index=False), start=4):
    ws5.cell(row=i, column=1, value=int(r.annee_surv))
    ws5.cell(row=i, column=2, value=int(r.nb_sin))
    ws5.cell(row=i, column=3, value=float(r.cout_1))
    ws5.cell(row=i, column=4, value=float(r.cout_2))
    ws5.cell(row=i, column=5, value=float(r.prime))
    ws5.cell(row=i, column=6, value=f"=C{i}/E{i}")
    ws5.cell(row=i, column=7, value=f"=D{i}/E{i}")
    autoformat_money(ws5.cell(row=i, column=3))
    autoformat_money(ws5.cell(row=i, column=4))
    autoformat_money(ws5.cell(row=i, column=5))
    autoformat_pct(ws5.cell(row=i, column=6), decimals=3)
    autoformat_pct(ws5.cell(row=i, column=7), decimals=3)
    for col in range(1, 8):
        ws5.cell(row=i, column=col).border = BORDER

# Total
n_last = 3 + len(df_bc1)
total_row = n_last + 1
ws5.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
ws5.cell(row=total_row, column=2, value=f"=SUM(B4:B{n_last})")
ws5.cell(row=total_row, column=3, value=f"=SUM(C4:C{n_last})")
ws5.cell(row=total_row, column=4, value=f"=SUM(D4:D{n_last})")
ws5.cell(row=total_row, column=5, value=f"=SUM(E4:E{n_last})")
ws5.cell(row=total_row, column=6, value=f"=C{total_row}/E{total_row}")
ws5.cell(row=total_row, column=7, value=f"=D{total_row}/E{total_row}")
for col in range(1, 8):
    ws5.cell(row=total_row, column=col).font = TOTAL_FONT
    ws5.cell(row=total_row, column=col).fill = TOTAL_FILL
    ws5.cell(row=total_row, column=col).border = BORDER
autoformat_money(ws5.cell(row=total_row, column=3))
autoformat_money(ws5.cell(row=total_row, column=4))
autoformat_money(ws5.cell(row=total_row, column=5))
autoformat_pct(ws5.cell(row=total_row, column=6), decimals=4)
autoformat_pct(ws5.cell(row=total_row, column=7), decimals=4)

# Prime pure 2024
prime_2024_row = total_row + 2
ws5.cell(row=prime_2024_row, column=1, value="Prime estimée 2024 (€)").font = Font(bold=True)
ws5.cell(row=prime_2024_row, column=2, value=prime_2024)
autoformat_money(ws5.cell(row=prime_2024_row, column=2))
ws5.cell(row=prime_2024_row+1, column=1, value="Prime pure Couche 1 (€)").font = Font(bold=True)
ws5.cell(row=prime_2024_row+1, column=2, value=f"=F{total_row}*B{prime_2024_row}")
autoformat_money(ws5.cell(row=prime_2024_row+1, column=2))
ws5.cell(row=prime_2024_row+2, column=1, value="Prime pure Couche 2 (€)").font = Font(bold=True)
ws5.cell(row=prime_2024_row+2, column=2, value=f"=G{total_row}*B{prime_2024_row}")
autoformat_money(ws5.cell(row=prime_2024_row+2, column=2))

for c in "ABCDEFG":
    ws5.column_dimensions[c].width = 18
ws5.column_dimensions["A"].width = 22

# ---- 6. BC AVEC CLAUSE ------------------------------------------------------
ws6 = wb.create_sheet("BC_avec_clause")
set_titre(ws6, 1, "Burning Cost AVEC clause de stabilité (indexation)", span=8)
indice = pd.read_csv(OUT_DIR / "indice_inflation.csv")
indice.columns = ["annee", "indice"]
set_header(ws6, 3, ["Année", "Indice (base 100 en 2014)"])
for i, r in enumerate(indice.itertuples(index=False), start=4):
    ws6.cell(row=i, column=1, value=int(r.annee))
    ws6.cell(row=i, column=2, value=float(r.indice))
    ws6.cell(row=i, column=2).number_format = "0.0000"
    for col in [1, 2]:
        ws6.cell(row=i, column=col).border = BORDER

# Méthode B : indexation par paiement
df_bc_b = pd.read_csv(OUT_DIR / "bc_avec_clause_methode_B.csv")
n_idx = 3 + len(indice)
start_bc = n_idx + 3
ws6.cell(row=start_bc-1, column=1, value="Méthode B : Indexation par paiement [REFERENCE]").font = Font(bold=True, size=12)
set_header(ws6, start_bc, ["Année surv.", "Nb sin.", "Coût Couche 1 indexé (€)",
                            "Coût Couche 2 indexé (€)", "Prime indexée (€)",
                            "BC Couche 1", "BC Couche 2"])
for i, r in enumerate(df_bc_b.itertuples(index=False), start=start_bc+1):
    ws6.cell(row=i, column=1, value=int(r.annee_surv))
    ws6.cell(row=i, column=2, value=int(r.nb_sin))
    ws6.cell(row=i, column=3, value=float(r.cout_1))
    ws6.cell(row=i, column=4, value=float(r.cout_2))
    ws6.cell(row=i, column=5, value=float(r.prime_indexee))
    ws6.cell(row=i, column=6, value=f"=C{i}/E{i}")
    ws6.cell(row=i, column=7, value=f"=D{i}/E{i}")
    autoformat_money(ws6.cell(row=i, column=3))
    autoformat_money(ws6.cell(row=i, column=4))
    autoformat_money(ws6.cell(row=i, column=5))
    autoformat_pct(ws6.cell(row=i, column=6), decimals=3)
    autoformat_pct(ws6.cell(row=i, column=7), decimals=3)
    for col in range(1, 8):
        ws6.cell(row=i, column=col).border = BORDER

n_last2 = start_bc + len(df_bc_b)
total_row2 = n_last2 + 1
ws6.cell(row=total_row2, column=1, value="TOTAL").font = TOTAL_FONT
ws6.cell(row=total_row2, column=2, value=f"=SUM(B{start_bc+1}:B{n_last2})")
ws6.cell(row=total_row2, column=3, value=f"=SUM(C{start_bc+1}:C{n_last2})")
ws6.cell(row=total_row2, column=4, value=f"=SUM(D{start_bc+1}:D{n_last2})")
ws6.cell(row=total_row2, column=5, value=f"=SUM(E{start_bc+1}:E{n_last2})")
ws6.cell(row=total_row2, column=6, value=f"=C{total_row2}/E{total_row2}")
ws6.cell(row=total_row2, column=7, value=f"=D{total_row2}/E{total_row2}")
for col in range(1, 8):
    ws6.cell(row=total_row2, column=col).font = TOTAL_FONT
    ws6.cell(row=total_row2, column=col).fill = TOTAL_FILL
    ws6.cell(row=total_row2, column=col).border = BORDER
autoformat_money(ws6.cell(row=total_row2, column=3))
autoformat_money(ws6.cell(row=total_row2, column=4))
autoformat_money(ws6.cell(row=total_row2, column=5))
autoformat_pct(ws6.cell(row=total_row2, column=6), decimals=4)
autoformat_pct(ws6.cell(row=total_row2, column=7), decimals=4)

# Prime pure 2024 (méthode B)
prime_row = total_row2 + 2
ws6.cell(row=prime_row, column=1, value="Prime estimée 2024 (€)").font = Font(bold=True)
ws6.cell(row=prime_row, column=2, value=prime_2024)
autoformat_money(ws6.cell(row=prime_row, column=2))
ws6.cell(row=prime_row+1, column=1, value="Prime pure Couche 1 (€)").font = Font(bold=True)
ws6.cell(row=prime_row+1, column=2, value=f"=F{total_row2}*B{prime_row}")
autoformat_money(ws6.cell(row=prime_row+1, column=2))
ws6.cell(row=prime_row+2, column=1, value="Prime pure Couche 2 (€)").font = Font(bold=True)
ws6.cell(row=prime_row+2, column=2, value=f"=G{total_row2}*B{prime_row}")
autoformat_money(ws6.cell(row=prime_row+2, column=2))

ws6.column_dimensions["A"].width = 22
for c in "BCDEFG":
    ws6.column_dimensions[c].width = 20

# ---- 7. PARETO --------------------------------------------------------------
ws7 = wb.create_sheet("Pareto_GPD")
set_titre(ws7, 1, "Modélisation Pareto / GPD de la sévérité", span=4)
set_header(ws7, 3, ["Paramètre", "Valeur"])
items = [
    ("Seuil de modélisation", "2 000 000 €"),
    ("Nombre d'excédents (>2M)", int(synth_pareto["n_excedents"])),
    ("Shape c (GPD)", round(synth_pareto["shape_c"], 4)),
    ("Scale sigma (GPD)", round(synth_pareto["scale_sigma"], 0)),
    ("Fréquence (sinistres/an)", round(synth_pareto["frequence_an"], 3)),
    ("E[coût Couche 1 | X > 2M] (M€)", round(synth_pareto["E_couche_1_par_sin_M"], 3)),
    ("E[coût Couche 2 | X > 2M] (M€)", round(synth_pareto["E_couche_2_par_sin_M"], 3)),
    ("Charge annuelle Couche 1 (M€)", round(synth_pareto["charge_annuelle_1_M"], 3)),
    ("Charge annuelle Couche 2 (M€)", round(synth_pareto["charge_annuelle_2_M"], 3)),
    ("Prime pure 2024 Couche 1 (M€)", round(synth_pareto["prime_pure_1_2024_M"], 3)),
    ("Prime pure 2024 Couche 2 (M€)", round(synth_pareto["prime_pure_2_2024_M"], 3)),
]
for i, (k, v) in enumerate(items, start=4):
    ws7.cell(row=i, column=1, value=k)
    ws7.cell(row=i, column=2, value=v)
    for col in [1, 2]:
        ws7.cell(row=i, column=col).border = BORDER
ws7.column_dimensions["A"].width = 35
ws7.column_dimensions["B"].width = 22

# ---- 8. PRIME COMMERCIALE ---------------------------------------------------
ws8 = wb.create_sheet("Prime_commerciale")
set_titre(ws8, 1, "Construction de la prime commerciale", span=4)

set_header(ws8, 3, ["Poste", "Couche 1 (€)", "Couche 2 (€)", "Commentaire"])
pp1 = prime_comm["prime_pure_couche_1"]
pp2 = prime_comm["prime_pure_couche_2"]
ch1 = prime_comm["k_ecart_type"] * prime_comm["sigma_couche_1"]
ch2 = prime_comm["k_ecart_type"] * prime_comm["sigma_couche_2"]

postes = [
    ("Prime pure", pp1, pp2, "BC indexé × prime 2024"),
    (f"+ Chargement sécurité (k={prime_comm['k_ecart_type']}, k×σ)",
     ch1, ch2, "Volatilité historique"),
    ("Prime technique", "=B4+B5", "=C4+C5", "Couvre la moyenne et la volatilité"),
    (f"+ Coût du capital ({int(prime_comm['cout_capital']*100)}% prime technique)",
     "=B6*0.06", "=C6*0.06", "Proxy du SCR alloué"),
    ("Prime brute", "=B6+B7", "=C6+C7", ""),
    (f"/ (1 - frais {int(prime_comm['frais_gestion']*100)}%)",
     "=B8/(1-0.05)", "=C8/(1-0.05)", "Frais de gestion réassureur"),
    ("PRIME COMMERCIALE", "=B9", "=C9", "Recommandation finale"),
]
for i, (poste, c1, c2, com) in enumerate(postes, start=4):
    ws8.cell(row=i, column=1, value=poste)
    ws8.cell(row=i, column=2, value=c1)
    ws8.cell(row=i, column=3, value=c2)
    ws8.cell(row=i, column=4, value=com)
    autoformat_money(ws8.cell(row=i, column=2))
    autoformat_money(ws8.cell(row=i, column=3))
    for col in range(1, 5):
        ws8.cell(row=i, column=col).border = BORDER
    if "PRIME" in poste or "technique" in poste or "brute" in poste:
        for col in range(1, 5):
            ws8.cell(row=i, column=col).font = TOTAL_FONT
            ws8.cell(row=i, column=col).fill = TOTAL_FILL

# Taux de prime
ws8.cell(row=12, column=1, value="Prime émise estimée 2024").font = Font(bold=True)
ws8.cell(row=12, column=2, value=prime_2024)
autoformat_money(ws8.cell(row=12, column=2))
ws8.cell(row=13, column=1, value="Taux de prime commerciale Couche 1").font = Font(bold=True)
ws8.cell(row=13, column=2, value=f"=B10/B12")
autoformat_pct(ws8.cell(row=13, column=2), decimals=3)
ws8.cell(row=14, column=1, value="Taux de prime commerciale Couche 2").font = Font(bold=True)
ws8.cell(row=14, column=2, value=f"=C10/B12")
autoformat_pct(ws8.cell(row=14, column=2), decimals=3)

ws8.column_dimensions["A"].width = 42
ws8.column_dimensions["B"].width = 22
ws8.column_dimensions["C"].width = 22
ws8.column_dimensions["D"].width = 40

# ---- Sauvegarde -------------------------------------------------------------
wb.save(XLSX_OUT)
print(f"Classeur Excel sauvegardé : {XLSX_OUT}")
